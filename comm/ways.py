# -*- coding: utf-8 -*-
__author__ = "wowo"
import pandas as pd
import math
import os
import json
import datetime
from calendar import calendar, month
from openpyxl import load_workbook
from openpyxl.styles import *
from zhdate import ZhDate
import calendar 
import excel2img
import smtplib
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import parseaddr, formataddr
import pymysql

__file__="../config.json"






class Way_a():
    def __init__(self):   
        # 系统配置
        with open(__file__,'r',encoding='utf-8') as config_data:
            self.configs=json.load(config_data) 
        self.now = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        # get database config
        localhost = self.configs['database_config']['localhost']
        username = self.configs['database_config']['username']
        password = self.configs['database_config']['password']
        database = self.configs['database_config']['database']
        # mysql connect
        self.conn = pymysql.connect(host='localhost',user='root',password='1001',database='wfjoa')
        self.hours = 22
        self.file = '../report/时段销售上报%s.xlsx'% self.now
        # 今年昨日
        x=['星期一','星期二','星期三','星期四','星期五','星期六','星期日']
        year=int(datetime.datetime.now().strftime('%Y'))
        month=int(datetime.datetime.now().strftime('%m'))
        day=int(datetime.datetime.now().strftime('%d')) 
        last_year_yesterday=int(day)-1
        if month==1 and last_year_yesterday==0:
            last_year=year-1
            last_year_month=12
            last_year_yesterday=calendar.monthrange(last_year,last_year_month)[-1]
        elif last_year_yesterday==0:
            last_year=year
            last_year_month=month-1
            last_year_yesterday=calendar.monthrange(last_year,last_year_month)[-1]
        else:
            last_year=year
            last_year_month=month
        self.yesterday_date=datetime.date(last_year,last_year_month,last_year_yesterday)
        zhdate=ZhDate.from_datetime(datetime.datetime(last_year,last_year_month,last_year_yesterday)).chinese()
        self.yesterday_weekday=x[datetime.date(last_year,last_year_month,last_year_yesterday).weekday()]
        self.yesterday_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
        # 去年昨天
        last_year=int(year)-1
        last_year_yesterday=int(day)-1
        if month==1 and last_year_yesterday==0:
            last_year=last_year-1
            last_year_month=12
            last_year_yesterday=calendar.monthrange(last_year,last_year_month)[-1]
        elif last_year_yesterday==0:
            last_year=last_year
            last_year_month=month-1
            last_year_yesterday=calendar.monthrange(last_year,last_year_month)[-1]
        else:
            last_year=last_year
            last_year_month=month
        self.last_yesterday_date=datetime.date(last_year,last_year_month,last_year_yesterday)
        zhdate=ZhDate.from_datetime(datetime.datetime(last_year,last_year_month,last_year_yesterday)).chinese()
        self.last_yesterday_weekday=x[datetime.date(last_year,last_year_month,last_year_yesterday).weekday()]
        self.last_yesterday_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
        # 去年今日
        last_year=int(year)-1
        self.last_before_yesterday_date=datetime.date(last_year,month,day)
        zhdate=ZhDate.from_datetime(datetime.datetime(last_year,month,day)).chinese()
        self.last_before_yesterday_weekday=x[datetime.date(last_year,month,day).weekday()]
        self.last_before_yesterday_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
    # 系统配置
    def config(self):
        with open(__file__,'r',encoding='utf-8') as config_data:
            configs=json.load(config_data)
            return configs
    # sql insert 
    def sql_insert(self,table,data):
        cur = self.conn.cursor()
        try:
            # create table
            sql="create table %s(%s varchar(40) primary key)" % (table,'insert_date')
            cur.execute(sql)
            # for add col
            for i in data:
                sql="alter table %s add %s varchar(120)" % (table,i)  
                cur.execute(sql) 
        except:
            pass
        finally:
            now=datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            for i,n in data.items():
                # date desc select first col data whether equal now_date
                sql="select insert_date from %s order by insert_date DESC limit 0,1" % (table)
                cur.execute(sql)
                x=cur.fetchall()
                # exist update data
                if x and str(x[0][0])==str(now):
                    sql="update %s set %s='%s' where insert_date='%s'" % (table,i,n,now)
                    cur.execute(sql)
                # inexistence insert data
                else:
                    sql="insert into %s(insert_date,%s) values('%s','%s')" % (table,i,now,n)
                    cur.execute(sql)
        cur.close()
        self.conn.commit()
        self.conn.close()
    # sql select 
    def sql_select(self,table,time_x):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where time='%s' order by insert_date DESC limit 0,1" % (table,time_x)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def sql_select_login_user(self,table,area):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where area='%s' order by insert_date DESC limit 0,1" % (table,area)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def sql_select_qq_user(self,table):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where smtp='smtp' order by insert_date DESC limit 0,1" % (table)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def create_excel(self):  
        configs=Way_a().config()    
        table_name_a=configs['database_config']['period_of_time_sale_table']
        self.yesterday_ly=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_ly"%(self.yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.yesterday_zl=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_zl"%(self.yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 去年今日
        self.last_yesterday_ly=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_ly"%(self.last_yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.last_yesterday_zl=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_zl"%(self.last_yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 去年明天
        self.last_before_yesterday_ly=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_ly"%(self.last_before_yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.last_before_yesterday_zl=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_zl"%(self.last_before_yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 每日毛利
        self.today_profit=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_rml"%(self.yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.tomonth_profit=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_yml"%(self.yesterday_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 每日客流
        self.yesterday_kl=Way_a().sql_select(table_name_a,"%s_%s_kl"%(self.yesterday_date,self.hours))["sale"]
        self.last_yesterday_kl=Way_a().sql_select(table_name_a,"%s_%s_kl"%(self.last_yesterday_date,self.hours))["sale"]
        self.last_before_yesterday_kl=Way_a().sql_select(table_name_a,"%s_%s_kl"%(self.last_before_yesterday_date,self.hours))["sale"]
        data=[  {'A':'',
                'B':self.yesterday_date,
                'C':self.yesterday_weekday,
                'D':self.last_yesterday_date,
                'E':self.last_yesterday_weekday,
                'F':self.last_before_yesterday_date,
                'G':self.last_before_yesterday_weekday},
                {'A':'',
                'B':self.yesterday_date_chinese,
                'C':'',
                'D':self.last_yesterday_date_chinese,
                'E':'',
                'F':self.last_before_yesterday_date_chinese,
                'G':''
                },
                {'A':'经营区域',
                'B':'总金额(万)',
                'C':'',
                'D':'总金额(万)',
                'E':'',
                'F':'总金额(万)',
                'G':''
                },
                {'A':'百货',
                'B':self.yesterday_ly,
                'C':'',
                'D':self.last_yesterday_ly,
                'E':'',
                'F':self.last_before_yesterday_ly,
                'G':''
                },
                {'A':'购物中心',
                'B':self.yesterday_zl,
                'C':'',
                'D':self.last_yesterday_zl,
                'E':'',
                'F':self.last_before_yesterday_zl,
                'G':''
                },
                {'A':'合计',
                'B':sum([self.yesterday_ly,self.yesterday_zl]),
                'C':'',
                'D':sum([self.last_yesterday_ly,self.last_yesterday_zl]),
                'E':'',
                'F':sum([self.last_before_yesterday_ly,self.last_before_yesterday_zl]),
                'G':''
                },
                {'A':'总销售',
                'B':'',
                'C':'',
                'D':'',
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'',
                'B':'',
                'C':'',
                'D':'',
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'进场客流数据',
                'B':self.yesterday_kl,
                'C':'',
                'D':self.last_yesterday_kl,
                'E':'',
                'F':self.last_before_yesterday_kl,
                'G':''
                },
                {'A':'',
                'B':'',
                'C':'',
                'D':'',
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'毛利',
                'B':'联营日毛利（万）',
                'C':'',
                'D':self.today_profit,
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'',
                'B':'联营月毛利（万）',
                'C':'',
                'D':self.tomonth_profit,
                'E':'',
                'F':'',
                'G':''
                }]
        data[0]['A']='截至%s:00' % self.hours
        data[1]['C']='截至%s:00' % self.hours
        data[1]['E']='截至%s:00' % self.hours
        data[1]['G']='截至%s:00' % self.hours
        df=pd.DataFrame(data)
        df.to_excel(self.file,sheet_name='sheet1',index=False,header=False)
        # 表格样式
    # def excel_style(self):
        wb=load_workbook(filename=self.file)
        sheet1=wb['sheet1']
        #调整列宽
        sheet1.column_dimensions['A'].width = 13
        sheet1.column_dimensions['B'].width = 13
        sheet1.column_dimensions['C'].width = 13
        sheet1.column_dimensions['D'].width = 13
        sheet1.column_dimensions['E'].width = 13
        sheet1.column_dimensions['F'].width = 13
        sheet1.column_dimensions['G'].width = 13
        #调整行高
        sheet1.row_dimensions[1].height = 18
        sheet1.row_dimensions[2].height = 18
        for i in range(3,8):
                sheet1.row_dimensions[i].height = 30
        sheet1.row_dimensions[9].height = 30
        for i in range(11,13):
                sheet1.row_dimensions[i].height = 30
        # 合并
        sheet1.merge_cells('A1:A2')
        #
        sheet1.merge_cells('B3:C3')
        sheet1.merge_cells('D3:E3')
        sheet1.merge_cells('F3:G3')
        #
        sheet1.merge_cells('B4:C4')
        sheet1.merge_cells('D4:E4')
        sheet1.merge_cells('F4:G4')
        #
        sheet1.merge_cells('B5:C5')
        sheet1.merge_cells('D5:E5')
        sheet1.merge_cells('F5:G5')
        #
        sheet1.merge_cells('B6:C6')
        sheet1.merge_cells('D6:E6')
        sheet1.merge_cells('F6:G6')
        #
        sheet1.merge_cells('B7:C7')
        sheet1.merge_cells('D7:E7')
        sheet1.merge_cells('F7:G7')
        #
        sheet1.merge_cells('B9:C9')
        sheet1.merge_cells('D9:E9')
        sheet1.merge_cells('F9:G9')
        #
        sheet1.merge_cells('B11:C11')
        sheet1.merge_cells('B12:C12')
        #
        sheet1.merge_cells('D11:E11')
        sheet1.merge_cells('D12:E12')
        #
        sheet1.merge_cells('A11:A12')
        # 设置居中，自动换行
        for i in range(1,13):
                sheet1['A%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['B%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['C%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['D%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['E%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['F%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['G%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        #边框
        border = Border(left=Side(border_style='thin',color='000000'),
                       right=Side(border_style='thin',color='000000'),
                       top=Side(border_style='thin',color='000000'),
                       bottom=Side(border_style='thin',color='000000'))
        for i in range(1,8):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
            sheet1['F%s'%i].border=border
            sheet1['G%s'%i].border=border
        for i in range(9,10):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
            sheet1['F%s'%i].border=border
            sheet1['G%s'%i].border=border
        for i in range(11,13):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
        #数值
        for i in range(4,8):
            sheet1['B%s'%i].number_format = '0.00'
            sheet1['C%s'%i].number_format = '0.00'
            sheet1['D%s'%i].number_format = '0.00'
            sheet1['E%s'%i].number_format = '0.00'
            sheet1['F%s'%i].number_format = '0.00'
            sheet1['G%s'%i].number_format = '0.00'
        for i in range(11,13):
            sheet1['D%s'%i].number_format = '0.00'
            sheet1['E%s'%i].number_format = '0.00'
        wb.save(self.file)
        #截图
        img_name=self.now
        sheet_list=["sheet1"]
        #接下来开始运行程序
        img_save = "../img/"
        if not os.path.exists(img_save):
            os.makedirs(img_save)
        # 保存为图片。
        try:
            print("开始截图，请耐心等待。。。")
            for i in range(len(sheet_list)):
                excel2img.export_img(self.file, img_save+img_name+".png", sheet_list[i], None)
        except :
            print("【没有截图成功！！！】请检查excel文件路径名称、工作表的名称是否全部正确！！！")
        else:
            print("截图成功，一共截图"+str(len(sheet_list))+"张图片。【保存在"+img_save+"】")
            

            
class Way_b():
    def __init__(self):
        # 系统配置
        with open(__file__,'r',encoding='utf-8') as config_data:
            self.configs=json.load(config_data) 
        # get database config
        localhost = self.configs['database_config']['localhost']
        username = self.configs['database_config']['username']
        password = self.configs['database_config']['password']
        database = self.configs['database_config']['database']
        # mysql connect
        self.conn = pymysql.connect(host='localhost',user='root',password='1001',database='wfjoa')
        self.now = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        self.hours=datetime.datetime.now().strftime('%H')
        self.file = '../report/时段销售上报%s.xlsx'% self.now
        # 今年今日
        x=['星期一','星期二','星期三','星期四','星期五','星期六','星期日']
        year=int(datetime.datetime.now().strftime('%Y'))
        month=int(datetime.datetime.now().strftime('%m'))
        day=int(datetime.datetime.now().strftime('%d')) 
        self.today_date=datetime.date(year,month,day)
        zhdate=ZhDate.from_datetime(datetime.datetime(year,month,day)).chinese()
        self.today_weekday=x[datetime.date(year,month,day).weekday()]
        self.today_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
        # 去年今日
        last_year_today=int(year)-1
        self.last_year_today_date=datetime.date(last_year_today,month,day)
        zhdate=ZhDate.from_datetime(datetime.datetime(last_year_today,month,day)).chinese()
        self.last_year_today_weekday=x[datetime.date(last_year_today,month,day).weekday()]
        self.last_year_today_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
        # 去年明日
        last_day_tomorrow=int(day)+1
        if month==12 and last_day_tomorrow-1==calendar.monthrange(last_year_today,month)[-1]:
            last_year_today=last_year_today+1
            last_year_month=1
            last_day_tomorrow=1
        elif last_day_tomorrow-1==calendar.monthrange(last_year_today,month)[-1]:
            last_year_month=month+1
            last_day_tomorrow=1
        else:
            last_year_today=last_year_today
            last_year_month=month
            last_day_tomorrow=last_day_tomorrow
        zhdate=ZhDate.from_datetime(datetime.datetime(last_year_today,last_year_month,last_day_tomorrow)).chinese()
        self.last_year_tomorrow_weekday=x[datetime.date(last_year_today,last_year_month,last_day_tomorrow).weekday()]
        self.last_year_tomorrow_date=datetime.date(last_year_today,last_year_month,last_day_tomorrow)
        self.last_year_tomorrow_date_chinese=((zhdate.split(' '))[0].split('年'))[-1]
    # 系统配置
    def config(self):
        with open(__file__,'r',encoding='utf-8') as config_data:
            configs=json.load(config_data)
            return configs
    # sql insert 
    def sql_insert(self,table,data):
        cur = self.conn.cursor()
        try:
            # create table
            sql="create table %s(%s varchar(40) primary key)" % (table,'insert_date')
            cur.execute(sql)
            # for add col
            for i in data:
                sql="alter table %s add %s varchar(120)" % (table,i)  
                cur.execute(sql) 
        except:
            pass
        finally:
            now=datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            for i,n in data.items():
                # date desc select first col data whether equal now_date
                sql="select insert_date from %s order by insert_date DESC limit 0,1" % (table)
                cur.execute(sql)
                x=cur.fetchall()
                # exist update data
                if x and str(x[0][0])==str(now):
                    sql="update %s set %s='%s' where insert_date='%s'" % (table,i,n,now)
                    cur.execute(sql)
                # inexistence insert data
                else:
                    sql="insert into %s(insert_date,%s) values('%s','%s')" % (table,i,now,n)
                    cur.execute(sql)
        cur.close()
        self.conn.commit()
        self.conn.close()
    # sql select 
    def sql_select(self,table,time_x):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where time='%s' order by insert_date DESC limit 0,1" % (table,time_x)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def sql_select_login_user(self,table,area):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where area='%s' order by insert_date DESC limit 0,1" % (table,area)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def sql_select_qq_user(self,table):
        cur = self.conn.cursor()
        data={}
        # date desc select first row
        sql="select * from %s where smtp='smtp' order by insert_date DESC limit 0,1" % (table)
        cur.execute(sql)
        select_data=cur.fetchall()
        # pymysql select col_naames
        sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s'" % table
        cur.execute(sql)
        row_name=cur.fetchall()
        for i in range(len(row_name)):
            # pymysql
            data['%s' % row_name[i][0]]=select_data[0][i]
        cur.close()
        self.conn.close()
        return data
    def create_excel(self):  
        configs=Way_b().config()    
        table_name_a=configs['database_config']['period_of_time_sale_table']
        # 今年今日
        self.today_ly=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_ly"%(self.today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.today_zl=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_zl"%(self.today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 去年今日
        self.last_year_today_ly=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_ly"%(self.last_year_today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.last_year_today_zl=round(float(str(Way_a().sql_select(table_name_a,"%s_%s_zl"%(self.last_year_today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 去年明天
        self.last_year_tomorrow_ly=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_ly"%(self.last_year_tomorrow_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.last_year_tomorrow_zl=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_zl"%(self.last_year_tomorrow_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 每日毛利
        self.today_profit=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_rml"%(self.today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        self.tomonth_profit=round(float(str(Way_b().sql_select(table_name_a,"%s_%s_yml"%(self.today_date,self.hours))["sale"]).replace(',',''))/10000,2)
        # 每日客流
        self.today_kl=Way_b().sql_select(table_name_a,"%s_%s_kl"%(self.today_date,self.hours))["sale"]
        self.last_year_today_kl=Way_b().sql_select(table_name_a,"%s_%s_kl"%(self.last_year_today_date,self.hours))["sale"]
        self.last_year_tomorrow_kl=Way_b().sql_select(table_name_a,"%s_%s_kl"%(self.last_year_tomorrow_date,self.hours))["sale"]
        data=[  {'A':'',
                'B':self.today_date,
                'C':self.today_weekday,
                'D':self.last_year_today_date,
                'E':self.last_year_today_weekday,
                'F':self.last_year_tomorrow_date,
                'G':self.last_year_tomorrow_weekday},
                {'A':'',
                'B':self.today_date_chinese,
                'C':'',
                'D':self.last_year_today_date_chinese,
                'E':'',
                'F':self.last_year_tomorrow_date_chinese,
                'G':''
                },
                {'A':'经营区域',
                'B':'总金额(万)',
                'C':'',
                'D':'总金额(万)',
                'E':'',
                'F':'总金额(万)',
                'G':''
                },
                {'A':'百货',
                'B':self.today_ly,
                'C':'',
                'D':self.last_year_today_ly,
                'E':'',
                'F':self.last_year_tomorrow_ly,
                'G':''
                },
                {'A':'购物中心',
                'B':self.today_zl,
                'C':'',
                'D':self.last_year_today_zl,
                'E':'',
                'F':self.last_year_tomorrow_zl,
                'G':''
                },
                {'A':'合计',
                'B':sum([self.today_ly,self.today_zl]),
                'C':'',
                'D':sum([self.last_year_today_ly,self.last_year_today_zl]),
                'E':'',
                'F':sum([self.last_year_tomorrow_ly,self.last_year_tomorrow_zl]),
                'G':''
                },
                {'A':'',
                'B':'',
                'C':'',
                'D':'',
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'进场客流数据',
                'B':self.today_kl,
                'C':'',
                'D':self.last_year_today_kl,
                'E':'',
                'F':self.last_year_tomorrow_kl,
                'G':''
                },
                {'A':'',
                'B':'',
                'C':'',
                'D':'',
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'毛利',
                'B':'联营日毛利（万）',
                'C':'',
                'D':self.today_profit,
                'E':'',
                'F':'',
                'G':''
                },
                {'A':'',
                'B':'联营月毛利（万）',
                'C':'',
                'D':self.tomonth_profit,
                'E':'',
                'F':'',
                'G':''
                }]
        data[0]['A']='截至%s:00' % self.hours
        data[1]['C']='截至%s:00' % self.hours
        data[1]['E']='截至%s:00' % self.hours
        data[1]['G']='截至%s:00' % self.hours
        df=pd.DataFrame(data)
        df.to_excel(self.file,sheet_name='sheet1',index=False,header=False)
        # 表格样式
    # def excel_style(self):
        wb=load_workbook(filename=self.file)
        sheet1=wb['sheet1']
        #调整列宽
        sheet1.column_dimensions['A'].width = 13
        sheet1.column_dimensions['B'].width = 13
        sheet1.column_dimensions['C'].width = 13
        sheet1.column_dimensions['D'].width = 13
        sheet1.column_dimensions['E'].width = 13
        sheet1.column_dimensions['F'].width = 13
        sheet1.column_dimensions['G'].width = 13
        #调整行高
        sheet1.row_dimensions[1].height = 18
        sheet1.row_dimensions[2].height = 18
        for i in range(3,7):
                sheet1.row_dimensions[i].height = 30
        sheet1.row_dimensions[8].height = 30
        for i in range(10,12):
                sheet1.row_dimensions[i].height = 30
        # 合并
        sheet1.merge_cells('A1:A2')
        #
        sheet1.merge_cells('B3:C3')
        sheet1.merge_cells('D3:E3')
        sheet1.merge_cells('F3:G3')
        #
        sheet1.merge_cells('B4:C4')
        sheet1.merge_cells('D4:E4')
        sheet1.merge_cells('F4:G4')
        #
        sheet1.merge_cells('B5:C5')
        sheet1.merge_cells('D5:E5')
        sheet1.merge_cells('F5:G5')
        #
        sheet1.merge_cells('B6:C6')
        sheet1.merge_cells('D6:E6')
        sheet1.merge_cells('F6:G6')
        #
        sheet1.merge_cells('B8:C8')
        sheet1.merge_cells('D8:E8')
        sheet1.merge_cells('F8:G8')
        #
        sheet1.merge_cells('B10:C10')
        sheet1.merge_cells('B11:C11')
        #
        sheet1.merge_cells('D10:E10')
        sheet1.merge_cells('D11:E11')
        #
        sheet1.merge_cells('A10:A11')
        # 设置居中，自动换行
        for i in range(1,12):
                sheet1['A%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['B%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['C%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['D%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['E%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['F%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                sheet1['G%s' % i ].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        #边框
        border = Border(left=Side(border_style='thin',color='000000'),
                       right=Side(border_style='thin',color='000000'),
                       top=Side(border_style='thin',color='000000'),
                       bottom=Side(border_style='thin',color='000000'))
        for i in range(1,7):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
            sheet1['F%s'%i].border=border
            sheet1['G%s'%i].border=border
        for i in range(8,9):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
            sheet1['F%s'%i].border=border
            sheet1['G%s'%i].border=border
        for i in range(10,12):
            sheet1['A%s'%i].border=border
            sheet1['B%s'%i].border=border
            sheet1['C%s'%i].border=border
            sheet1['D%s'%i].border=border
            sheet1['E%s'%i].border=border
        #数值
        for i in range(4,7):
            sheet1['B%s'%i].number_format = '0.00'
            sheet1['C%s'%i].number_format = '0.00'
            sheet1['D%s'%i].number_format = '0.00'
            sheet1['E%s'%i].number_format = '0.00'
            sheet1['F%s'%i].number_format = '0.00'
            sheet1['G%s'%i].number_format = '0.00'
        for i in range(10,12):
            sheet1['D%s'%i].number_format = '0.00'
            sheet1['E%s'%i].number_format = '0.00'
        wb.save(self.file)
        #截图
        img_name=self.now
        sheet_list=["sheet1"]
        #接下来开始运行程序
        img_save = "../img/"
        if not os.path.exists(img_save):
            os.makedirs(img_save)
        # 保存为图片。
        try:
            print("开始截图，请耐心等待。。。")
            for i in range(len(sheet_list)):
                excel2img.export_img(self.file, img_save+img_name+".png", sheet_list[i], None)
        except :
            print("【没有截图成功！！！】请检查excel文件路径名称、工作表的名称是否全部正确！！！")
        else:
            print("截图成功，一共截图"+str(len(sheet_list))+"张图片。【保存在"+img_save+"】")
        

    




class Smtp():
    def __init__(self):
        # 获取报表
        excel_name='../report/'
        base_dir=excel_name
        l=os.listdir(excel_name)
        l.sort(key=lambda fn: os.path.getmtime(base_dir+fn) if not os.path.isdir(base_dir+fn) else 0)
        self.excel_ip=excel_name+l[-1]
        print(self.excel_ip)
        # 获取截图
        img_name='../img/'
        base_dir=img_name
        l=os.listdir(img_name)
        l.sort(key=lambda fn: os.path.getmtime(base_dir+fn) if not os.path.isdir(base_dir+fn) else 0)
        self.img_ip=img_name+l[-1]
    def format_addr(self,x):
        name, addr = parseaddr(x)
        return formataddr((Header(name, 'utf-8').encode(), addr))
    def send_mail_excel(self):
        qq_user=Way_a().sql_select_qq_user("qq_email")
        login_user=qq_user["login_user"]
        login_authorization_code=qq_user["login_code"]
        receive_user=qq_user["receive_user"]
        msg = MIMEMultipart()
        msg.attach(MIMEText('send with file...', 'plain', 'utf-8'))
        msg['From'] = Smtp().format_addr('徐铃纹 <%s>' % login_user)
        msg['To'] = Smtp().format_addr('管理员 <%s>' % receive_user)
        msg['Subject'] = Header('时段销售数据', 'utf-8').encode()
        att1 = MIMEText(open('%s' % self.excel_ip, 'rb').read(), 'base64', 'utf-8')
        att1["Content-Type"] = 'application/octet-stream'
        att1["Content-Disposition"] = 'attachment; filename= %s' %  self.excel_ip
        msg.attach(att1)
        stmp_server='smtp.qq.com'
        server=smtplib.SMTP(stmp_server,25)
        server.set_debuglevel(1)
        server.login(login_user,login_authorization_code)
        server.sendmail(login_user,[receive_user],msg.as_string())
        server.quit()
    def send_mail_img(self):
        qq_user=Way_a().sql_select_qq_user("qq_email")
        login_user=qq_user["login_user"]
        login_authorization_code=qq_user["login_code"]
        receive_user=qq_user["receive_user"]
        msg = MIMEMultipart()
        msg.attach(MIMEText('send with file...', 'plain', 'utf-8'))
        msg['From'] = Smtp().format_addr('徐铃纹 <%s>' % login_user)
        msg['To'] = Smtp().format_addr('管理员 <%s>' % receive_user)
        msg['Subject'] = Header('时段销售数据', 'utf-8').encode()
        att1 = MIMEText(open('%s' % self.img_ip, 'rb').read(), 'base64', 'utf-8')
        att1["Content-Type"] = 'application/octet-stream'
        att1["Content-Disposition"] = 'attachment; filename= %s' %  self.img_ip
        msg.attach(att1)
        stmp_server='smtp.qq.com'
        server=smtplib.SMTP(stmp_server,25)
        server.set_debuglevel(1)
        server.login(login_user,login_authorization_code)
        server.sendmail(login_user,[receive_user],msg.as_string())
        server.quit()



class DataBase_user_add():
    #存入QQ邮箱到数据库
    def database_qq_email(self):
        login_user=input("请输入发送方邮箱：")
        login_code=input("请输入发送方邮箱授权码：")
        receive_user=input("请输入接收方邮箱：")
        data={      "smtp":"smtp",
                    "login_user":login_user,
                    "login_code":login_code,
                    "receive_user":receive_user
                    }
        Way_a().sql_insert("qq_email",data)
        print("存储成功")



    #存入用户信息到数据库
    def database_user(self):
        url_oa=input("请输入wfj_oa系统的地址：")
        username_oa=input("请输入wfj_oa系统的登录账号：")
        password_oa=input("请输入wfj_oa系统的登录密码：")
        data_oa={  "area":"oa",
                "url":url_oa,
                "username":username_oa,
                "password":password_oa
                }
        url_kl=input("请输入wfj_kl系统的地址：")
        username_kl=input("请输入wfj_kl系统的登录账号：")   
        password_kl=input("请输入wfj_kl系统的登录密码：")
        data_kl={  "area":"kl",
                "url":url_kl,
                "username":username_kl,
                "password":password_kl
                }
        Way_a().sql_insert("wfj_user",data_oa)
        Way_a().sql_insert("wfj_user",data_kl)
        print("存储成功")
        

